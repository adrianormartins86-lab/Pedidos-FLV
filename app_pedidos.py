import streamlit as st
import pandas as pd
import io
import streamlit.components.v1 as components
from streamlit_gsheets import GSheetsConnection
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# CONFIGURAÇÃO DA PÁGINA
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="Gestão de Pedidos - FLV Normal",
    page_icon="🛒",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ─────────────────────────────────────────────
# CSS GLOBAL E DE IMPRESSÃO
# ─────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600&family=IBM+Plex+Sans:wght@400;500;700&display=swap');

:root {
    --bg-main:        #0d1117;
    --bg-card:        #161b22;
    --bg-sidebar:     #0d1117;
    --green-dark:     #1a3a2a;
    --green-mid:      #1f4d35;
    --green-accent:   #2ea043;
    --green-bright:   #3fb950;
    --green-glow:     rgba(46,160,67,.25);
    --text-primary:   #e6edf3;
    --text-muted:     #7d8590;
    --text-header:    #cae8cb;
    --border:         #21262d;
    --border-active:  #2ea043;
    --row-hover:      rgba(46,160,67,.08);
    --row-selected:   rgba(46,160,67,.18);
}

.stApp, .main { background-color: var(--bg-main) !important; color: var(--text-primary) !important; }
html, body, [class*="css"] { font-family: 'IBM Plex Sans', sans-serif !important; }
section[data-testid="stSidebar"] { background-color: var(--bg-sidebar) !important; border-right: 1px solid var(--border); }
section[data-testid="stSidebar"] * { color: var(--text-primary) !important; }
section[data-testid="stSidebar"] .stRadio label { font-size: 14px; }

.stButton > button[kind="primary"] {
    background: linear-gradient(135deg, var(--green-mid) 0%, var(--green-accent) 100%) !important;
    color: #fff !important;
    border: 1px solid var(--green-accent) !important;
    border-radius: 8px !important;
    font-weight: 700 !important;
    letter-spacing: .3px;
    transition: all .2s ease !important;
}
.stButton > button[kind="primary"]:hover {
    transform: translateY(-2px) !important;
    box-shadow: 0 4px 18px var(--green-glow) !important;
    border-color: var(--green-bright) !important;
}
.stButton > button {
    background: var(--bg-card) !important;
    color: var(--text-primary) !important;
    border: 1px solid var(--border) !important;
    border-radius: 8px !important;
    transition: all .2s ease !important;
}
.stButton > button:hover {
    border-color: var(--green-accent) !important;
    color: var(--green-bright) !important;
    transform: translateY(-1px) !important;
}
.stTextInput input, .stSelectbox > div > div {
    background-color: var(--bg-card) !important;
    border: 1px solid var(--border) !important;
    border-radius: 8px !important;
    color: var(--text-primary) !important;
}
.stTextInput input:focus, .stSelectbox > div > div:focus-within {
    border-color: var(--green-accent) !important;
    box-shadow: 0 0 0 3px var(--green-glow) !important;
}
.title-input input {
    font-weight: 700 !important;
    font-size: 16px !important;
    color: var(--green-bright) !important;
    padding: 2px 8px !important;
    background: transparent !important;
    border: 1px dashed #21262d !important;
}
.title-input input:focus { border: 1px dashed #2ea043 !important; }

[data-testid="stDataEditor"] [data-testid="glideDataEditor"] .gdg-header-cell,
[data-testid="stDataEditor"] .dvn-stack .gdg-header {
    background-color: var(--green-dark) !important;
    color: var(--text-header) !important;
}
[data-testid="stDataEditor"] {
    border-radius: 10px !important;
    overflow: hidden;
    border: 1px solid var(--green-mid) !important;
    box-shadow: 0 4px 20px rgba(0,0,0,.4);
    font-size: 12px !important; 
}
[data-testid="stDataEditor"] .gdg-cell.gdg-selected,
[data-testid="stDataEditor"] .gdg-cell[data-state="focused"],
[data-testid="stDataEditor"] .gdg-cell[aria-selected="true"] {
    background-color: var(--row-selected) !important;
    outline: 2px solid var(--green-accent) !important;
    outline-offset: -2px;
}
[data-testid="stDataEditor"] .gdg-row:hover .gdg-cell { background-color: var(--row-hover) !important; }

div[data-testid="stVerticalBlockBorderWrapper"] {
    background-color: var(--bg-card) !important;
    border: 1px solid var(--border) !important;
    border-radius: 12px !important;
    transition: box-shadow .25s ease, border-color .25s ease;
}
div[data-testid="stVerticalBlockBorderWrapper"]:hover {
    border-color: var(--green-mid) !important;
    box-shadow: 0 6px 24px rgba(0,0,0,.35) !important;
}
[data-testid="stMetric"] {
    background-color: var(--bg-card);
    border: 1px solid var(--border);
    border-radius: 10px;
    padding: 10px 10px;
}
[data-testid="stMetricValue"] { color: var(--green-bright) !important; font-weight: 700; font-size: 1.8rem !important; }
[data-testid="stMetricLabel"] { color: var(--text-muted) !important; }

.sidebar-hidden section[data-testid="stSidebar"],
.sidebar-hidden [data-testid="collapsedControl"] { display: none !important; }
.sidebar-hidden .main .block-container { max-width: 100% !important; padding-left: 2rem !important; padding-right: 2rem !important; }

.topbar-loja {
    background: linear-gradient(90deg, var(--green-dark) 0%, #0d2018 100%);
    border: 1px solid var(--green-mid);
    border-radius: 10px;
    padding: 10px 18px;
    margin-bottom: 18px;
    display: flex;
    align-items: center;
    justify-content: space-between;
}
.topbar-left { display: flex; align-items: center; gap: 12px; }
.topbar-title { font-size: 18px; font-weight: 700; color: var(--text-header); }
.topbar-sub { font-size: 11px; color: var(--text-muted); margin-top: 2px; }

@media print {
    @page { margin: 10mm; }
    .stApp, .main, body, html {
        background-color: #ffffff !important;
        color: #000000 !important;
        background-image: none !important;
        padding: 0 !important;
        margin: 0 !important;
    }
    .main .block-container, [data-testid="stAppViewBlockContainer"] {
        padding-top: 0 !important;
        margin-top: 0 !important;
    }
    header, [data-testid="stSidebar"], [data-testid="stHeader"] { display: none !important; }
    [data-testid="stElementContainer"]:has([data-testid="stDataEditor"]),
    [data-testid="stElementContainer"]:has(.topbar-loja),
    [data-testid="stElementContainer"]:has([data-testid="stMetric"]),
    [data-testid="stElementContainer"]:has(button),
    [data-testid="stHorizontalBlock"],
    div[data-testid="stVerticalBlockBorderWrapper"],
    hr, .stAlert, .stInfo { display: none !important; }
    #print-section {
        display: block !important;
        width: 100% !important;
        margin-top: 0 !important;
        padding-top: 0 !important;
    }
    #print-section h2 {
        font-size: 16px !important;
        margin: 0 0 10px 0 !important;
        padding-bottom: 5px !important;
        border-bottom: 1px solid #000 !important;
        color: #000 !important;
    }
    .print-container { display: flex; gap: 15px; align-items: flex-start; width: 100%; }
    .print-col { flex: 1; width: 50%; }
    table.print-table {
        width: 100%;
        border-collapse: collapse;
        font-size: 10px !important;
        color: #000000 !important;
        font-family: 'IBM Plex Sans', sans-serif;
        line-height: 1.05 !important;
    }
    table.print-table th, table.print-table td {
        border: 1px solid #000000 !important;
        padding: 2px 4px !important;
        text-align: left;
        color: #000000 !important;
        background-color: #ffffff !important;
    }
    table.print-table th {
        background-color: #e0e0e0 !important;
        font-weight: bold;
        -webkit-print-color-adjust: exact;
        print-color-adjust: exact;
    }
    table.print-table tr { break-inside: avoid !important; page-break-inside: avoid !important; }
}
@media screen {
    #print-section { display: none !important; }
}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# CONSTANTES E DADOS INICIAIS
# ─────────────────────────────────────────────
LOJAS = ["Loja 01", "Loja 02", "Loja 03", "Loja 04", "Loja 05", "Loja 06", "Loja 07", "Loja 08"]
NOVOS_NOMES_LOJAS = ["291", "292", "293", "294", "295", "296", "297", "298"]
MAPA_LOJAS = dict(zip(LOJAS, NOVOS_NOMES_LOJAS))
FORNECEDORES_ESPECIAIS_LINHA = ["BANANA SANTOME", "MELANCIA CARLIN", "MELANCIA MARCINHO", "RODRIGO BATATA"]

produtos_iniciais = [
    {"Código": 1571, "Descrição": "Abacate Cx 20 Kg", "Tipo": "Box"},
    {"Código": 2614, "Descrição": "Abacaxi Doce Mel Cx c/7", "Tipo": "Box"},
    {"Código": 95, "Descrição": "Abacaxi Hawai Un", "Tipo": "Pedra"},
    {"Código": 94, "Descrição": "Abacaxi Perola Un", "Tipo": "Box"},
    {"Código": 232, "Descrição": "Abobora Cabotia 20 Kg", "Tipo": "Box"},
    {"Código": 235, "Descrição": "Abobora Gigante Doce kg", "Tipo": "Box"},
    {"Código": 236, "Descrição": "Abobora Italia Bdj", "Tipo": "Pedra"},
    {"Código": 45, "Descrição": "Abobora Italia Cx 20 Kg", "Tipo": "Pedra"},
    {"Código": 237, "Descrição": "Abobora Menina Bdj", "Tipo": "Pedra"},
    {"Código": 56, "Descrição": "Abobora Menina Cx 20Kg", "Tipo": "Pedra"},
    {"Código": 238, "Descrição": "Abobora Moranga Saco 20kg", "Tipo": "Box"},
    {"Código": 240, "Descrição": "Abobora Paulista Verde cx 22Kg", "Tipo": "Pedra"},
    {"Código": 85, "Descrição": "Acelga Cx c/8", "Tipo": "Pedra"},
    {"Código": 1746, "Descrição": "Alface Americana unid", "Tipo": "Pedra"},
    {"Código": 9001, "Descrição": "Alho Nacional Cx 10Kg", "Tipo": "Box"},
    {"Código": 320, "Descrição": "Alho Poro dz", "Tipo": "Box"},
    {"Código": 894, "Descrição": "Ameixa Importada Cx 9Kg", "Tipo": "Box"},
    {"Código": 9002, "Descrição": "Ameixa Nacional", "Tipo": "Box"},
    {"Código": 504, "Descrição": "Amendoim sc 10Kg", "Tipo": "Box"},
    {"Código": 113, "Descrição": "Aspargos", "Tipo": "Box"},
    {"Código": 896, "Descrição": "Atemoia Cx 4Kg", "Tipo": "Box"},
    {"Código": 897, "Descrição": "Avocado Cx 10Kg", "Tipo": "Box"},
    {"Código": 2567, "Descrição": "Banana Maça Cx", "Tipo": "Pedra"},
    {"Código": 2568, "Descrição": "Banana Nanica Cx", "Tipo": "Pedra"},
    {"Código": 2569, "Descrição": "Banana Prata Cx", "Tipo": "Pedra"},
    {"Código": 98, "Descrição": "Banana Terra Cx 20Kg", "Tipo": "Box"},
    {"Código": 551, "Descrição": "Batata Asterix Saq 25Kg", "Tipo": "Pedra"},
    {"Código": 73, "Descrição": "Batata Doce Branca Cx 22Kg", "Tipo": "Pedra"},
    {"Código": 60, "Descrição": "Batata Doce Cx 22Kg", "Tipo": "Pedra"},
    {"Código": 508, "Descrição": "Batata kg Saq 25Kg", "Tipo": "Pedra"},
    {"Código": 26, "Descrição": "Batata Yacom Kg", "Tipo": "Box"},
    {"Código": 61, "Descrição": "Berinjela Cx 13Kg", "Tipo": "Pedra"},
    {"Código": 2732, "Descrição": "Berinjela Japonesa Bdja", "Tipo": "Pedra"},
    {"Código": 62, "Descrição": "Berinjela Japonesa Cx 13Kg", "Tipo": "Pedra"},
    {"Código": 256, "Descrição": "Beterraba Cx 21kg", "Tipo": "Pedra"},
    {"Código": 64, "Descrição": "Brocolis Chines BDJ", "Tipo": "Box"},
    {"Código": 707, "Descrição": "Cabotia 300g Bjda Descascada", "Tipo": "Pedra"},
    {"Código": 28, "Descrição": "Caju bandeija cx c/4", "Tipo": "Box"},
    {"Código": 9003, "Descrição": "Caqui Fuyu cx 20Kg", "Tipo": "Box"},
    {"Código": 9004, "Descrição": "Caqui Kioto / Chocolate cx 20Kg", "Tipo": "Box"},
    {"Código": 264, "Descrição": "Caqui Rama Forte Cx 5Kg", "Tipo": "Box"},
    {"Código": 69, "Descrição": "Cara Cx 22Kg", "Tipo": "Pedra"},
    {"Código": 127, "Descrição": "Carambola bandeija cx c/4", "Tipo": "Box"},
    {"Código": 74, "Descrição": "Caxi Cx 20Kg", "Tipo": "Pedra"},
    {"Código": 2730, "Descrição": "Cebola Branca Bdja", "Tipo": "Pedra"},
    {"Código": 9000, "Descrição": "Cebola cx 3 Saco 20Kg", "Tipo": "Box"},
    {"Código": 43, "Descrição": "Cebola Roxa Saco 20Kg", "Tipo": "Box"},
    {"Código": 17, "Descrição": "Cenoura Baby un", "Tipo": "Box"},
    {"Código": 267, "Descrição": "Cenoura Cx 21kg", "Tipo": "Pedra"},
    {"Código": 19, "Descrição": "Champignon Paris 250G un", "Tipo": "Box"},
    {"Código": 902, "Descrição": "Chuchu Cx 20Kg", "Tipo": "Box"},
    {"Código": 1555, "Descrição": "Cidra Ralada Pré Cozid Un", "Tipo": "Box"},
    {"Código": 21, "Descrição": "Coco Seco Cx 18Kg", "Tipo": "Box"},
    {"Código": 1700, "Descrição": "Coco Verde saco c/10", "Tipo": "Box"},
    {"Código": 87, "Descrição": "Couve Flor Bdj", "Tipo": "Box"},
    {"Código": 86, "Descrição": "Couve Flor dz", "Tipo": "Pedra"},
    {"Código": 108, "Descrição": "Ervilha em Grãos Bdja", "Tipo": "Pedra"},
    {"Código": 109, "Descrição": "Ervilha Horta Torta Bdj", "Tipo": "Pedra"},
    {"Código": 279, "Descrição": "Figo Pre Cozido un", "Tipo": "Box"},
    {"Código": 128, "Descrição": "Figo Roxo bandeija cx c/3un", "Tipo": "Box"},
    {"Código": 712, "Descrição": "Gengibre Cx 12Kg", "Tipo": "Box"},
    {"Código": 281, "Descrição": "Gobo Un", "Tipo": "Box"},
    {"Código": 59, "Descrição": "Goiaba Cx 20Kg", "Tipo": "Box"},
    {"Código": 1662, "Descrição": "Inhame Cx 22kg", "Tipo": "Pedra"},
    {"Código": 42, "Descrição": "Jilo Bdj", "Tipo": "Pedra"},
    {"Código": 41, "Descrição": "Jilo Cx 15Kg", "Tipo": "Pedra"},
    {"Código": 112, "Descrição": "Kiwi 500g Bdj", "Tipo": "Box"},
    {"Código": 904, "Descrição": "Kiwi Importado Cx t23 cx 9Kg", "Tipo": "Box"},
    {"Código": 1651, "Descrição": "Laranja Bahia Cx 18Kg", "Tipo": "Box"},
    {"Código": 1599, "Descrição": "Laranja Bahia importada Cx 15Kg", "Tipo": "Box"},
    {"Código": 1307, "Descrição": "Laranja Lima Cx 18Kg", "Tipo": "Box"},
    {"Código": 9005, "Descrição": "Laranja Lima PC 1,5KG", "Tipo": "Pedra"},
    {"Código": 1516, "Descrição": "Laranja P/ Suco Cx 20Kg", "Tipo": "Pedra"},
    {"Código": 53, "Descrição": "Laranja Pera Cx 20Kg", "Tipo": "Pedra"},
    {"Código": 9006, "Descrição": "Laranja Pera PC 3 KG", "Tipo": "Pedra"},
    {"Código": 288, "Descrição": "Laranja Pre Cozida un", "Tipo": "Box"},
    {"Código": 13, "Descrição": "Lima Da Persia Cx 10Kg", "Tipo": "Box"},
    {"Código": 44, "Descrição": "Limao Cx 22kg 22Kg", "Tipo": "Box"},
    {"Código": 581, "Descrição": "Limão Rosa Bdja", "Tipo": "Pedra"},
    {"Código": 91, "Descrição": "Limao Rosa Cx 20Kg", "Tipo": "Pedra"},
    {"Código": 522, "Descrição": "Limao Siciliano Cx 15Kg", "Tipo": "Box"},
    {"Código": 291, "Descrição": "Maça Argentina Cx 18Kg", "Tipo": "Box"},
    {"Código": 9007, "Descrição": "Maça Fuji Cx 18Kg", "Tipo": "Box"},
    {"Código": 9008, "Descrição": "Maça Gala Cx 18Kg", "Tipo": "Box"},
    {"Código": 1697, "Descrição": "Maça Gransmith Cx 1/2 9Kg", "Tipo": "Box"},
    {"Código": 1652, "Descrição": "Maça Pacote 1kg diversos cx 18un", "Tipo": "Box"},
    {"Código": 2052, "Descrição": "Maça Pink Lady Cx 18Kg", "Tipo": "Box"},
    {"Código": 106, "Descrição": "Mamao Formosa Cx 10Kg", "Tipo": "Box"},
    {"Código": 3, "Descrição": "Mamao Papaya Cx 10kg 15un", "Tipo": "Box"},
    {"Código": 75, "Descrição": "Mandioca Desc 1Kg", "Tipo": "Pedra"},
    {"Código": 78, "Descrição": "Mandioca Salsa Bdj", "Tipo": "Pedra"},
    {"Código": 76, "Descrição": "Mandioca Salsa Cx 10Kg", "Tipo": "Pedra"},
    {"Código": 406, "Descrição": "Manga Espada 6Kg", "Tipo": "Box"},
    {"Código": 6, "Descrição": "Manga Palmer Cx 18Kg", "Tipo": "Box"},
    {"Código": 130, "Descrição": "Manga Rosa 18Kg", "Tipo": "Box"},
    {"Código": 908, "Descrição": "Manga Tomy Cx 19kg", "Tipo": "Box"},
    {"Código": 92, "Descrição": "Maracuja Azedo Cx 10Kg", "Tipo": "Box"},
    {"Código": 1646, "Descrição": "Maracuja Doce Cx Plastica 10Kg", "Tipo": "Box"},
    {"Código": 518, "Descrição": "Maxi Pecan 250G un", "Tipo": "Box"},
    {"Código": 546, "Descrição": "Maxixe Bandeja 300g", "Tipo": "Pedra"},
    {"Código": 3003, "Descrição": "Melancia Amarela", "Tipo": "Pedra"},
    {"Código": 673, "Descrição": "Melancia Baby Cx 14Kg", "Tipo": "Box"},
    {"Código": 2, "Descrição": "Melancia Favo de Mel", "Tipo": "Pedra"},
    {"Código": 1, "Descrição": "Melancia Un", "Tipo": "Pedra"},
    {"Código": 9009, "Descrição": "Melão Amarelo Gaia Cx 13kg", "Tipo": "Box"},
    {"Código": 1409, "Descrição": "Melão Bebezinho", "Tipo": "Box"},
    {"Código": 198, "Descrição": "Melao Cantalupe Cx 10Kg", "Tipo": "Box"},
    {"Código": 412, "Descrição": "Melao Cepi Amarelo Cx 10Kg", "Tipo": "Box"},
    {"Código": 200, "Descrição": "Melao Dino Cx 10Kg", "Tipo": "Box"},
    {"Código": 202, "Descrição": "Melao Galia Cx 10Kg", "Tipo": "Box"},
    {"Código": 1407, "Descrição": "Melao Orange Cx 6Kg", "Tipo": "Box"},
    {"Código": 424, "Descrição": "Melao Rei Cx 10Kg", "Tipo": "Box"},
    {"Código": 206, "Descrição": "Melao Rei Sapo Cx 10Kg", "Tipo": "Box"},
    {"Código": 915, "Descrição": "Melao Sapo Cx 13Kg", "Tipo": "Box"},
    {"Código": 477, "Descrição": "Mexerica cx 20kg", "Tipo": "Box"},
    {"Código": 72, "Descrição": "Milho Verde Bdj", "Tipo": "Pedra"},
    {"Código": 20, "Descrição": "Mirtilo cx c/12", "Tipo": "Box"},
    {"Código": 58, "Descrição": "Moranguinho Bdj cx c/4", "Tipo": "Pedra"},
    {"Código": 536, "Descrição": "Moricote Ole cx 18Kg", "Tipo": "Box"},
    {"Código": 209, "Descrição": "Moyashi un", "Tipo": "Box"},
    {"Código": 79, "Descrição": "Nabo maco c/6", "Tipo": "Pedra"},
    {"Código": 916, "Descrição": "Nectarina importada 9Kg", "Tipo": "Box"},
    {"Código": 22, "Descrição": "Nespera Bdj c/4", "Tipo": "Box"},
    {"Código": 57, "Descrição": "Pepino Caipira Cx 20Kg", "Tipo": "Pedra"},
    {"Código": 46, "Descrição": "Pepino Fuchinari Cx 20Kg", "Tipo": "Pedra"},
    {"Código": 2009, "Descrição": "Pera Argentina cx 18Kg", "Tipo": "Box"},
    {"Código": 118, "Descrição": "Pera asiatica ou Hossui cx 9Kg", "Tipo": "Box"},
    {"Código": 119, "Descrição": "Pera Erconini 1kg bdj cx c/10", "Tipo": "Box"},
    {"Código": 1431, "Descrição": "Pera Nacional cx 20Kg", "Tipo": "Box"},
    {"Código": 1600, "Descrição": "Pera Portuguesa bdj", "Tipo": "Box"},
    {"Código": 121, "Descrição": "Pera Portuguesa Cx 9Kg", "Tipo": "Box"},
    {"Código": 4, "Descrição": "Pera Red Cx 18Kg", "Tipo": "Box"},
    {"Código": 891, "Descrição": "Pessego importado cx 9Kg", "Tipo": "Box"},
    {"Código": 537, "Descrição": "Physalis cx c/8", "Tipo": "Box"},
    {"Código": 52, "Descrição": "Pimenta Americana Cx 10Kg", "Tipo": "Pedra"},
    {"Código": 80, "Descrição": "Pimenta Biquinho Bdj", "Tipo": "Pedra"},
    {"Código": 83, "Descrição": "Pimenta Gode Bdj", "Tipo": "Pedra"},
    {"Código": 540, "Descrição": "Pimenta Vermelha Bdj", "Tipo": "Pedra"},
    {"Código": 47, "Descrição": "Pimentao Amarelo Cx 10Kg", "Tipo": "Pedra"},
    {"Código": 949, "Descrição": "Pimentão Misto Bdj", "Tipo": "Pedra"},
    {"Código": 49, "Descrição": "Pimentao Verde Cx 10Kg", "Tipo": "Pedra"},
    {"Código": 48, "Descrição": "Pimentao Vermelho Cx 10Kg", "Tipo": "Pedra"},
    {"Código": 138, "Descrição": "Pinha cx 4,5kg", "Tipo": "Box"},
    {"Código": 498, "Descrição": "Pinhao saq 10Kg", "Tipo": "Box"},
    {"Código": 139, "Descrição": "Pitaia Cx 10kg", "Tipo": "Box"},
    {"Código": 1486, "Descrição": "Poncan cx 20Kg", "Tipo": "Box"},
    {"Código": 40, "Descrição": "Quiabo Bdj", "Tipo": "Pedra"},
    {"Código": 110, "Descrição": "Rabanete Bdj", "Tipo": "Pedra"},
    {"Código": 2886, "Descrição": "Rabanete Maco Dz 12un", "Tipo": "Pedra"},
    {"Código": 88, "Descrição": "Repolho dz 12un", "Tipo": "Pedra"},
    {"Código": 84, "Descrição": "Repolho Manteiga Bj dz 12un", "Tipo": "Pedra"},
    {"Código": 140, "Descrição": "Repolho Roxo Dz 12un", "Tipo": "Pedra"},
    {"Código": 16, "Descrição": "Roma cx 4,5kg", "Tipo": "Box"},
    {"Código": 31, "Descrição": "Salsao un", "Tipo": "Box"},
    {"Código": 32, "Descrição": "Shimeji Branco un", "Tipo": "Box"},
    {"Código": 679, "Descrição": "Shimeji Preto un", "Tipo": "Box"},
    {"Código": 33, "Descrição": "Shitake un", "Tipo": "Box"},
    {"Código": 10, "Descrição": "Tamara Bandeija Palito Un", "Tipo": "Box"},
    {"Código": 23, "Descrição": "Tamarindo bdj cx c/4", "Tipo": "Box"},
    {"Código": 364, "Descrição": "Tofu un", "Tipo": "Box"},
    {"Código": 39, "Descrição": "Tomate Saladete Cx 22Kg", "Tipo": "Pedra"},
    {"Código": 51, "Descrição": "Tomate Sweet Grape un", "Tipo": "Box"},
    {"Código": 538, "Descrição": "Tomatinho Bdj Naranti", "Tipo": "Pedra"},
    {"Código": 147, "Descrição": "Tomatinho cocktail holandez cx 6Kg", "Tipo": "Box"},
    {"Código": 100, "Descrição": "Uva Niagara Bdja cx c/10", "Tipo": "Box"},
    {"Código": 9010, "Descrição": "Uva Preta 500g Bdja cx c/10", "Tipo": "Box"},
    {"Código": 9011, "Descrição": "Uva Preta Benetaka Bdja cx c/10", "Tipo": "Box"},
    {"Código": 9012, "Descrição": "Uva Verde 500g Bdja cx c/10", "Tipo": "Box"},
    {"Código": 9013, "Descrição": "Uva Vermelha 500g Bdja cx c/10", "Tipo": "Box"},
    {"Código": 68, "Descrição": "Vagem Bdj", "Tipo": "Pedra"},
    {"Código": 67, "Descrição": "Vagem kg Cx 11kg", "Tipo": "Pedra"}
]

mapa_inicial_codigos = {
    "NIDE": [45, 49, 67, 57, 46, 48, 47], "Claudir Mendes": [57, 46, 49], "SANDRO": [75],
    "DENIZE": [45, 49, 67, 57, 46, 48, 47, 41, 52, 69, 56], "JOVANO": [1746, 88, 49, 140, 85],
    "JEFINHO": [85, 140, 256, 267, 88, 57, 45, 49, 1662, 46], "LUCIANO": [61, 41, 49, 56, 45],
    "THIAGO": [61, 91, 67, 74, 49, 52, 45, 56], "CRISTIAN": [40, 949, 42, 83, 68, 538, 78],
    "ROGERIO NARANTE": [538], "FERNANDO NARANTE": [46], "SILVIO MAND SALSA": [76],
    "HORTA": [108, 109], "GLAUCIA MACIEL": [84, 85], "ALEMÃO": [39], "RENAN SS": [72],
    "NEGUIN": [85, 86, 88, 61, 45], "RODRIGO CHANAN": [85, 86, 88, 61, 1662, 140],
    "MARCELO MORANGO": [58], "JOÃO BATISTA": [79, 60, 56, 1662, 69], "GIACOMELLO": [95],
    "PRIMO": [240, 86, 49, 45, 88, 85], "RENATO MANDIOCA": [75], "THIAGO SERRA": [91, 49, 45, 56, 61],
    "TICO": [236, 237, 707, 2730, 42, 581, 78, 546, 80, 83, 540, 949, 40, 110, 68, 109],
    "ALGACIR": [1516, 53], "MAURICIO": [62], "PAULO IGASHIBAHI": [47, 48],
    "GILSOM BATATA": [508, 551], "DORI BATATA": [508, 551], "BANANA SANTOME": [2567, 2569, 2568],
    "MELANCIA CARLIN": [1], "MELANCIA MARCINHO": [673, 1, 3003], "RODRIGO BATATA": [508]
}

# ─────────────────────────────────────────────
# CONEXÃO GOOGLE SHEETS & FUNÇÕES DE DADOS
# ─────────────────────────────────────────────
conn = st.connection("gsheets", type=GSheetsConnection)

@st.cache_data(ttl=15)
def carregar_banco():
    df_prod = conn.read(worksheet="Produtos")
    df_ped = conn.read(worksheet="Pedidos")
    df_est = conn.read(worksheet="Estoque")

    mudou_algo = False

    if df_prod.empty or "Código" not in df_prod.columns:
        df_prod = pd.DataFrame(produtos_iniciais)
        for loja in LOJAS: df_prod[loja] = True
        conn.update(worksheet="Produtos", data=df_prod)
        mudou_algo = True

    if df_ped.empty or "Código" not in df_ped.columns:
        df_ped = pd.DataFrame(columns=["Código"] + LOJAS + ["R$Preço", "OBS:"])
        df_ped["Código"] = df_prod["Código"]
        df_ped[LOJAS] = 0
        df_ped["R$Preço"] = 0.0
        df_ped["OBS:"] = ""
        conn.update(worksheet="Pedidos", data=df_ped)
        mudou_algo = True

    if df_est.empty or "Código" not in df_est.columns:
        df_est = pd.DataFrame(columns=["Código"] + LOJAS)
        df_est["Código"] = df_prod["Código"]
        df_est[LOJAS] = 0
        conn.update(worksheet="Estoque", data=df_est)
        mudou_algo = True

    novos_ped = df_prod[~df_prod["Código"].isin(df_ped["Código"])]["Código"]
    if not novos_ped.empty:
        df_n_ped = pd.DataFrame({"Código": novos_ped})
        df_n_ped[LOJAS] = 0
        df_n_ped["R$Preço"] = 0.0
        df_n_ped["OBS:"] = ""
        df_ped = pd.concat([df_ped, df_n_ped], ignore_index=True)
        conn.update(worksheet="Pedidos", data=df_ped)
        mudou_algo = True

    novos_est = df_prod[~df_prod["Código"].isin(df_est["Código"])]["Código"]
    if not novos_est.empty:
        df_n_est = pd.DataFrame({"Código": novos_est})
        df_n_est[LOJAS] = 0
        df_est = pd.concat([df_est, df_n_est], ignore_index=True)
        conn.update(worksheet="Estoque", data=df_est)
        mudou_algo = True

    for loja in LOJAS:
        if loja in df_ped.columns: df_ped[loja] = pd.to_numeric(df_ped[loja], errors='coerce').fillna(0).astype(int)
        if loja in df_est.columns: df_est[loja] = pd.to_numeric(df_est[loja], errors='coerce').fillna(0).astype(int)
        if loja in df_prod.columns: df_prod[loja] = df_prod[loja].fillna(False).astype(bool)

    if "R$Preço" in df_ped.columns: df_ped["R$Preço"] = pd.to_numeric(df_ped["R$Preço"], errors='coerce').fillna(0.0)
    if "OBS:" in df_ped.columns: df_ped["OBS:"] = df_ped["OBS:"].fillna("").astype(str)

    if mudou_algo:
        st.cache_data.clear()

    return df_prod, df_ped, df_est

df_produtos, df_pedidos, df_estoque = carregar_banco()

LISTA_NOMES_PRODUTOS = [str(x) for x in df_produtos['Descrição'].unique()]

lista_cfg = []
for f, cods in mapa_inicial_codigos.items():
    for c in cods:
        desc_match = df_produtos[df_produtos['Código'] == c]['Descrição']
        if not desc_match.empty:
            desc = desc_match.values[0]
            lista_cfg.append({"Fornecedor": f, "Produto": desc})
df_fornecedores_config = pd.DataFrame(lista_cfg)

if 'reset_counter' not in st.session_state:
    st.session_state['reset_counter'] = 0

# ─────────────────────────────────────────────
# SISTEMA DE LOGIN
# ─────────────────────────────────────────────
if 'usuario_logado' not in st.session_state:
    st.session_state['usuario_logado'] = None

if st.session_state['usuario_logado'] is None:
    st.write("<br><br>", unsafe_allow_html=True)
    _, col2, _ = st.columns([1, 1.4, 1])

    with col2:
        with st.container(border=True):
            h1, h2 = st.columns([4, 1])
            with h1:
                st.markdown("""
                    <h2 style='margin-bottom:0'>Portal de Pedidos</h2>
                    <p style='color:#7d8590;font-size:14px;margin-top:4px'>FLV Normal — Molicenter</p>
                """, unsafe_allow_html=True)
            with h2:
                st.write("")
                try:
                    st.image("passaro_logo.png", width=60)
                except Exception:
                    st.markdown("🐦", unsafe_allow_html=True)

            st.divider()

            usuarios_permitidos = ["Selecione..."] + ["Administrador"] + LOJAS
            usuario_selecionado = st.selectbox("👤 Usuário de acesso:", usuarios_permitidos)
            senha_digitada = st.text_input("🔑 Senha de acesso:", type="password", autocomplete="off")

            st.write("<br>", unsafe_allow_html=True)

            if st.button("Entrar no Sistema", type="primary", use_container_width=True):
                if usuario_selecionado == "Selecione...":
                    st.error("⚠️ Por favor, selecione um usuário.")
                elif usuario_selecionado == "Administrador" and senha_digitada == "moli0000":
                    st.session_state['usuario_logado'] = usuario_selecionado
                    st.rerun()
                elif usuario_selecionado in LOJAS and senha_digitada == "moli1234":
                    st.session_state['usuario_logado'] = usuario_selecionado
                    st.rerun()
                elif senha_digitada:
                    st.error("⚠️ Senha incorreta. Tente novamente.")

            st.markdown('<p style="font-size: 11px; color: #7d8590; text-align: center; margin-top: 10px;">🔒 Acesso restrito — Molicenter © 2026</p>', unsafe_allow_html=True)
    st.stop()

# ─────────────────────────────────────────────
# PÓS-LOGIN
# ─────────────────────────────────────────────
usuario_atual = st.session_state['usuario_logado']
acesso_total  = usuario_atual == "Administrador"

if not acesso_total:
    st.markdown("""
    <script>
        document.body.classList.add('sidebar-hidden');
        const root = window.parent.document.querySelector('.stApp');
        if (root) root.classList.add('sidebar-hidden');
    </script>
    <style>
        section[data-testid="stSidebar"] { display: none !important; }
        [data-testid="collapsedControl"]  { display: none !important; }
        .main .block-container {
            max-width: 100% !important;
            padding-left: 2.5rem !important;
            padding-right: 2.5rem !important;
        }
    </style>
    """, unsafe_allow_html=True)

# ─────────────────────────────────────────────
# MODAL ZERAR TUDO
# ─────────────────────────────────────────────
@st.dialog("🚨 Confirmação Necessária")
def modal_zerar_estoque_pedidos():
    st.markdown("Tem certeza que deseja **zerar todos os pedidos e o estoque** de todas as lojas?")
    st.markdown("⚠️ *Esta ação limpará também os preços e observações direto no Google Sheets e não poderá ser desfeita.*")
    
    st.write("<br>", unsafe_allow_html=True)
    c1, c2 = st.columns(2)
    with c1:
        if st.button("❌ Não, cancelar", use_container_width=True):
            st.rerun()
    with c2:
        if st.button("✔️ Sim, zerar tudo", type="primary", use_container_width=True):
            st.session_state['reset_counter'] += 1
            df_prod, df_ped, df_est = carregar_banco()
            
            df_ped[LOJAS] = 0
            df_ped["R$Preço"] = 0.0
            df_ped["OBS:"] = ""
            df_est[LOJAS] = 0
            
            conn.update(worksheet="Pedidos", data=df_ped)
            conn.update(worksheet="Estoque", data=df_est)
            st.cache_data.clear()
            st.rerun()

# ─────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────
with st.sidebar:
    try:
        st.image("passaro_logo.png", width=72)
    except Exception:
        st.markdown("🐦")

    st.markdown(f"### Olá, **{usuario_atual}**")
    st.caption("Sistema de Pedidos Integrado")
    st.divider()

    if acesso_total:
        perfil_navegacao = st.radio("📍 Navegação:", [
            "Separação e Fechamento",
            "Visão das Lojas",
            "Visão Fornecedores (Ademilto)",
            "Catálogo de Produtos"
        ])
    else:
        perfil_navegacao = "Visão das Lojas"

    st.divider()

    total_preenchidos = (df_pedidos[LOJAS] > 0).any(axis=1).sum()
    st.metric("Itens c/ pedido", total_preenchidos, help="Itens que têm ao menos 1 quantidade preenchida")

    st.divider()
    if st.button("🚪 Sair / Logout", use_container_width=True):
        st.session_state['usuario_logado'] = None
        st.rerun()

# ─────────────────────────────────────────────
# HELPER: gera o arquivo Excel formatado
# ─────────────────────────────────────────────
def _gerar_excel_formatado(df_editado_admin, filtro_setor):
    """
    Recebe o DataFrame já editado e o filtro de setor.
    Retorna bytes do .xlsx pronto para download.
    """
    HDR_BG    = "C55A11"   # laranja escuro — cabeçalho
    HDR_FG    = "FFFFFF"   # branco
    GREEN_ROW = "E2EFDA"   # verde claro alternado
    WHITE_ROW = "FFFFFF"   # branco alternado
    TOTAL_BG  = "C6EFCE"   # verde suave — coluna TOTAL
    PRICE_BG  = "FCE4D6"   # laranja suave — coluna PREÇO

    thin = Side(style="thin", color="BFBFBF")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    df_exp = df_editado_admin.copy()
    df_exp = df_exp.rename(columns=MAPA_LOJAS)

    if filtro_setor in ("Box", "Pedra"):
        df_exp = df_exp.rename(columns={
            "Código":      "CODIGO",
            "Descrição":   "PRODUTOS MOLICENTER",
            "TOTAL GERAL": "TOTAL",
            "R$Preço":     "PREÇO",
        })

    # Nomes das colunas de base (CODIGO/Código, PRODUTO/Descrição, [Tipo])
    cod_col  = "CODIGO"      if filtro_setor in ("Box", "Pedra") else "Código"
    prod_col = "PRODUTOS MOLICENTER" if filtro_setor in ("Box", "Pedra") else "Descrição"
    tot_col  = "TOTAL"       if filtro_setor in ("Box", "Pedra") else "TOTAL GERAL"
    pre_col  = "PREÇO"       if filtro_setor in ("Box", "Pedra") else "R$Preço"
    obs_col  = "OBS:"

    base_cols = [cod_col, prod_col]
    if "Tipo" in df_exp.columns:
        base_cols.append("Tipo")

    store_cols  = NOVOS_NOMES_LOJAS              # ["291"…"298"]
    spacer_cols = [" " * i for i in range(1, 7)] # 6 colunas vazias ocultas
    for s in spacer_cols:
        df_exp[s] = ""

    final_cols = base_cols + store_cols + spacer_cols + [tot_col, pre_col, obs_col]
    df_exp = df_exp[[c for c in final_cols if c in df_exp.columns]]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pedidos FLV"

    HEADER_ROW = 2    # linha 1 fica em branco; cabeçalho na linha 2
    DATA_START  = 3

    cols = list(df_exp.columns)

    # ── Cabeçalho ──────────────────────────────────────────────────────────
    for ci, col_name in enumerate(cols, 1):
        cell = ws.cell(row=HEADER_ROW, column=ci, value=col_name)
        cell.font      = Font(bold=True, color=HDR_FG, name="Arial", size=9)
        cell.fill      = PatternFill("solid", start_color=HDR_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = brd

    # AutoFiltro na linha de cabeçalho
    ws.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(len(cols))}{HEADER_ROW}"

    # ── Dados ──────────────────────────────────────────────────────────────
    for ri, (_, row) in enumerate(df_exp.iterrows(), DATA_START):
        row_bg = GREEN_ROW if (ri - DATA_START) % 2 == 0 else WHITE_ROW

        for ci, col_name in enumerate(cols, 1):
            raw = row[col_name]
            # Zera vira célula vazia
            if raw == 0 or raw == 0.0 or str(raw).strip() in ("0", "0.0", "nan", ""):
                raw = None

            cell = ws.cell(row=ri, column=ci, value=raw)
            cell.font   = Font(name="Arial", size=9)
            cell.border = brd

            if col_name in (prod_col, "Descrição", "PRODUTOS MOLICENTER"):
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

            if col_name == tot_col:
                cell.fill = PatternFill("solid", start_color=TOTAL_BG)
                cell.font = Font(name="Arial", size=9, bold=True)
            elif col_name == pre_col:
                cell.fill = PatternFill("solid", start_color=PRICE_BG)
                if raw is not None:
                    ccell.number_format = '[$R$-pt-BR] #,##0.00'
            else:
                cell.fill = PatternFill("solid", start_color=row_bg)

    # ── Larguras das colunas ────────────────────────────────────────────────
    widths = {
        cod_col:  8,
        prod_col: 34,
        "Tipo":   8,
        tot_col:  8,
        pre_col:  12,
        obs_col:  22,
    }
    for s in store_cols:
        widths[s] = 6
    for s in spacer_cols:
        widths[s] = 3

    for ci, col_name in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(col_name, 8)

    # ── Ocultar colunas de espaçamento (K:P equivalentes) ──────────────────
    for ci, col_name in enumerate(cols, 1):
        if col_name.strip() == "":
            ws.column_dimensions[get_column_letter(ci)].hidden = True

    # ── Congelar painéis: mantém CODIGO+PRODUTO visíveis ao rolar ──────────
    freeze_col = len(base_cols) + 1
    ws.freeze_panes = f"{get_column_letter(freeze_col)}{DATA_START}"

    # ── Alturas das linhas ──────────────────────────────────────────────────
    ws.row_dimensions[1].height          = 6   # linha 1 em branco
    ws.row_dimensions[HEADER_ROW].height = 18
    for ri in range(DATA_START, DATA_START + len(df_exp)):
        ws.row_dimensions[ri].height = 15

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ─────────────────────────────────────────────
# ROTA 1: SEPARAÇÃO E FECHAMENTO
# ─────────────────────────────────────────────
if perfil_navegacao == "Separação e Fechamento":
    st.markdown("""
    <div class="page-header" style="background: linear-gradient(90deg, var(--green-dark) 0%, #0d2018 100%); padding: 14px 20px; border-radius: 10px; margin-bottom: 22px;">
        <span style="font-size: 26px; margin-right: 12px;">📊</span>
        <div style="display: inline-block; vertical-align: top;">
            <div style="font-size: 20px; font-weight: 700; color: var(--text-header);">Separação e Fechamento</div>
            <div style="font-size: 12px; color: var(--text-muted); margin-top: 2px;">Consolidado geral de quantidades (O estoque das lojas não é exibido aqui)</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    with st.container(border=True):
        filtro_setor = st.radio("🔍 Filtrar Exibição por Setor:", ["Todos", "Box", "Pedra"], horizontal=True)
        st.write("<br>", unsafe_allow_html=True)

        df_base  = df_produtos[["Código","Descrição","Tipo"]]
        df_final = pd.merge(df_base, df_pedidos, on="Código")
        df_final["TOTAL GERAL"] = df_final[LOJAS].sum(axis=1)
        cols_order = ["Código", "Descrição", "Tipo"] + LOJAS + ["TOTAL GERAL", "R$Preço", "OBS:"]
        df_final = df_final[cols_order]

        if filtro_setor != "Todos":
            df_final = df_final[df_final["Tipo"] == filtro_setor].reset_index(drop=True)
            df_final = df_final.drop(columns=["Tipo"])

        col_cfg = {
            "Código":      st.column_config.NumberColumn(width=80, format="%d", disabled=True),
            "Descrição":   st.column_config.TextColumn(disabled=True),
            "TOTAL GERAL": st.column_config.NumberColumn("TOTAL ▶", width=90, format="%d", disabled=True),
            "R$Preço":     st.column_config.NumberColumn("R$ Preço", width=100, format="R$ %.2f", min_value=0.0, step=0.01),
            "OBS:":        st.column_config.TextColumn("OBS:", width=200)
        }
        
        if "Tipo" in df_final.columns:
            col_cfg["Tipo"] = st.column_config.TextColumn("Setor", width=100, disabled=True)
            
        for loja, novo_nome in MAPA_LOJAS.items():
            col_cfg[loja] = st.column_config.NumberColumn(novo_nome, format="%d", min_value=0, step=1)

        df_editado_admin = st.data_editor(
            df_final, hide_index=True, use_container_width=True,
            height=580, column_config=col_cfg,
            key=f"admin_editor_{st.session_state['reset_counter']}"
        )

        st.divider()
        col_salvar, col_csv, col_excel, col_limpa, _ = st.columns([2.5, 1.5, 1.5, 2, 2.5])

        with col_salvar:
            if st.button("💾 Salvar Alterações", type="primary", use_container_width=True):
                _, df_ped_fresco, _ = carregar_banco()
                for _, row in df_editado_admin.iterrows():
                    mask = df_ped_fresco["Código"] == row["Código"]
                    for loja in LOJAS:
                        df_ped_fresco.loc[mask, loja] = row[loja]
                    df_ped_fresco.loc[mask, "R$Preço"] = row["R$Preço"]
                    df_ped_fresco.loc[mask, "OBS:"] = row["OBS:"]
                
                conn.update(worksheet="Pedidos", data=df_ped_fresco)
                st.cache_data.clear()
                st.success("✅ Ajustes, preços e observações salvos com sucesso no Sheets!")
                st.rerun()

        with col_csv:
            df_csv = df_editado_admin.copy()
            df_csv = df_csv.rename(columns=MAPA_LOJAS)
            csv = df_csv.to_csv(index=False).encode("utf-8")
            st.download_button("⬇️ CSV", data=csv, file_name="separacao_semanal.csv", mime="text/csv", use_container_width=True)
            
        with col_excel:
            excel_bytes = _gerar_excel_formatado(df_editado_admin, filtro_setor)
            nome_arquivo_excel = f"separacao_{filtro_setor.lower()}.xlsx" if filtro_setor != "Todos" else "separacao_semanal.xlsx"
            st.download_button(
                "⬇️ Excel",
                data=excel_bytes,
                file_name=nome_arquivo_excel,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

        with col_limpa:
            if st.button("🚨 Zerar Pedidos/Estoque", use_container_width=True):
                modal_zerar_estoque_pedidos()

# ─────────────────────────────────────────────
# ROTA 2: VISÃO DAS LOJAS
# ─────────────────────────────────────────────
elif perfil_navegacao == "Visão das Lojas":
    if acesso_total:
        loja_selecionada = st.selectbox("👁️ Visualizar como:", LOJAS)
    else:
        loja_selecionada = usuario_atual

    col_info, col_logout = st.columns([8, 2])
    with col_info:
        id_loja = MAPA_LOJAS.get(loja_selecionada, loja_selecionada)
        st.markdown(f"""
        <div class="topbar-loja">
            <div class="topbar-left">
                <span style="font-size:22px">📋</span>
                <div>
                    <div class="topbar-title">{loja_selecionada} ({id_loja}) — FLV Normal</div>
                    <div class="topbar-sub">Preencha o estoque atual e a quantidade necessária para o pedido</div>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)
    with col_logout:
        st.markdown("<div style='height:14px'></div>", unsafe_allow_html=True)
        if st.button("🚪 Sair / Logout", use_container_width=True):
            st.session_state['usuario_logado'] = None
            st.rerun()

    df_visiveis = df_produtos[df_produtos[loja_selecionada] == True]
    df_loja = df_visiveis[["Código","Descrição","Tipo"]].copy()
    df_est = df_estoque[["Código", loja_selecionada]].rename(columns={loja_selecionada: "Estoque"})
    df_qtd = df_pedidos[["Código", loja_selecionada]].rename(columns={loja_selecionada: "Qtde"})
    df_loja = pd.merge(df_loja, df_est, on="Código", how="left")
    df_loja = pd.merge(df_loja, df_qtd, on="Código", how="left")

    with st.container(border=True):
        st.info("💡 **Dica:** Preencha primeiro o **Estoque** e depois a **Qtde** do pedido.")
        
        col_cfg_loja = {
            "Código":         st.column_config.NumberColumn(width=85, format="%d", disabled=True),
            "Descrição":      st.column_config.TextColumn(width=400, disabled=True),
            "Tipo":           st.column_config.TextColumn("Setor", width=100, disabled=True),
            "Estoque":        st.column_config.NumberColumn("📦 Estoque", width=120, min_value=0, step=1),
            "Qtde":           st.column_config.NumberColumn("🛒 Qtde", width=120, min_value=0, step=1)
        }
        
        _, col_tabela, _ = st.columns([1, 4, 1])
        
        with col_tabela:
            df_editado = st.data_editor(
                df_loja, column_config=col_cfg_loja,
                hide_index=True, use_container_width=True, height=520,
                key=f"loja_editor_{st.session_state['reset_counter']}"
            )

        df_imprimir = df_editado.copy()
        df_imprimir["Código"] = df_imprimir["Código"].fillna(0).astype(int).astype(str)
        df_imprimir = df_imprimir.rename(columns={"Tipo": "Setor", "Estoque": "Est.", "Qtde": "Ped."})
        
        meio = (len(df_imprimir) // 2) + (len(df_imprimir) % 2)
        df1 = df_imprimir.iloc[:meio]
        df2 = df_imprimir.iloc[meio:]
        
        html1 = df1.to_html(index=False, classes="print-table")
        html2 = df2.to_html(index=False, classes="print-table") if not df2.empty else ""
        
        st.markdown(f"""
        <div id="print-section">
            <h2 style="color: black; margin-bottom: 10px; text-align: center; border-bottom: 2px solid black; padding-bottom: 5px;">
                Resumo do Pedido — {loja_selecionada} ({id_loja})
            </h2>
            <div class="print-container">
                <div class="print-col">{html1}</div>
                <div class="print-col">{html2}</div>
            </div>
        </div>
        """, unsafe_allow_html=True)

        itens_com_pedido = int((df_editado["Qtde"] > 0).sum())
        total_itens      = len(df_editado)
        total_unidades   = int(df_editado["Qtde"].sum())
        pct              = round(itens_com_pedido / total_itens * 100) if total_itens else 0

        st.divider()
        m1, m2, m3, col_print, col_btn = st.columns([2.5, 2.2, 1.8, 1.5, 3])
        with m1: st.metric("Itens preenchidos (Pedido)", f"{itens_com_pedido} / {total_itens}")
        with m2: st.metric("Total de unidades", total_unidades)
        with m3: st.metric("Cobertura", f"{pct}%")
        
        with col_print:
            st.write("<br>", unsafe_allow_html=True)
            if st.button("🖨️ Imprimir", use_container_width=True):
                components.html("<script>window.parent.print();</script>", height=0)
                
        with col_btn:
            st.write("<br>", unsafe_allow_html=True)
            if st.button("💾 Salvar Pedido da Semana", type="primary", use_container_width=True):
                _, df_ped_fresco, df_est_fresco = carregar_banco()
                
                for _, row in df_editado.iterrows():
                    mask_ped = df_ped_fresco["Código"] == row["Código"]
                    mask_est = df_est_fresco["Código"] == row["Código"]
                    df_ped_fresco.loc[mask_ped, loja_selecionada] = row["Qtde"]
                    df_est_fresco.loc[mask_est, loja_selecionada] = row["Estoque"]
                
                conn.update(worksheet="Pedidos", data=df_ped_fresco)
                conn.update(worksheet="Estoque", data=df_est_fresco)
                st.cache_data.clear()
                st.success(f"✅ Estoque e Pedido da {loja_selecionada} salvos na nuvem!")

# ─────────────────────────────────────────────
# ROTA 3: VISÃO FORNECEDORES (ADEMILTO)
# ─────────────────────────────────────────────
elif perfil_navegacao == "Visão Fornecedores (Ademilto)":
    st.markdown("""
    <div class="page-header" style="background: linear-gradient(90deg, var(--green-dark) 0%, #0d2018 100%); padding: 14px 20px; border-radius: 10px; margin-bottom: 22px;">
        <span style="font-size: 26px; margin-right: 12px;">🚚</span>
        <div style="display: inline-block; vertical-align: top;">
            <div style="font-size: 20px; font-weight: 700; color: var(--text-header);">Visão Fornecedores (Ademilto) - Modo de Edição Livre</div>
            <div style="font-size: 12px; color: var(--text-muted); margin-top: 2px;">Altere nomes, códigos e totais livremente antes de gerar seu print. Adicione ou remova linhas na própria tabela.</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    df_base_produtos = df_produtos[["Código", "Descrição"]]
    df_base_pedidos = df_pedidos.copy()
    df_base_pedidos["Total"] = df_base_pedidos[LOJAS].sum(axis=1)

    df_consolidado = pd.merge(df_base_produtos, df_base_pedidos[["Código", "Total", "R$Preço"]], on="Código", how="inner")
    nomes_fornecedores = df_fornecedores_config["Fornecedor"].unique()
    
    for i in range(0, len(nomes_fornecedores), 2):
        cols = st.columns(2, gap="small")
        for j, fornecedor in enumerate(nomes_fornecedores[i:i+2]):
            descricoes_fornecedor = df_fornecedores_config[df_fornecedores_config["Fornecedor"] == fornecedor]["Produto"].tolist()
            codigos_do_fornecedor = df_base_produtos[df_base_produtos["Descrição"].isin(descricoes_fornecedor)]["Código"].tolist()
            
            with cols[j]:
                with st.container(border=True):
                    st.markdown('<div class="title-input">', unsafe_allow_html=True)
                    st.text_input("Fornecedor", value=f"🛒 {fornecedor}", label_visibility="collapsed", key=f"title_{fornecedor}_{st.session_state['reset_counter']}")
                    st.markdown('</div>', unsafe_allow_html=True)
                    
                    if fornecedor in FORNECEDORES_ESPECIAIS_LINHA:
                        df_ped_esp = df_base_pedidos[df_base_pedidos["Código"].isin(codigos_do_fornecedor)]
                        dict_lojas = {"Visão": LOJAS + ["TOTAL"]}
                        col_configs_especial = {"Visão": st.column_config.TextColumn("Visão", disabled=False)}
                        
                        for cod in codigos_do_fornecedor:
                            desc_series = df_base_produtos[df_base_produtos["Código"] == cod]["Descrição"]
                            desc = desc_series.values[0] if not desc_series.empty else "Prod"
                            
                            partes = desc.split()
                            palavra = " ".join(partes[:2]) if len(partes) > 1 else desc
                            nome_col = f"{cod} - {palavra}"
                            
                            valores_lojas = []
                            for loja in LOJAS:
                                val = df_ped_esp[df_ped_esp["Código"] == cod][loja].values
                                valores_lojas.append(int(val[0]) if len(val) > 0 else 0)
                            
                            valores_lojas.append(sum(valores_lojas))
                            dict_lojas[nome_col] = valores_lojas
                            col_configs_especial[nome_col] = st.column_config.NumberColumn(nome_col, format="%d", disabled=False)
                            
                        df_especial = pd.DataFrame(dict_lojas)
                        altura_esp = int((len(df_especial) + 2) * 36) + 5
                        
                        st.data_editor(
                            df_especial, 
                            hide_index=True, 
                            use_container_width=True, 
                            column_config=col_configs_especial,
                            height=altura_esp,
                            num_rows="dynamic",
                            key=f"forn_esp_{fornecedor}_{st.session_state['reset_counter']}"
                        )
                    
                    else:
                        df_fornecedor = df_consolidado[df_consolidado["Código"].isin(codigos_do_fornecedor)].copy()
                        df_fornecedor = df_fornecedor.rename(columns={"Código": "Cód", "Descrição": "Produtos", "R$Preço": "R$ Preço"})
                        df_fornecedor["R$ Total"] = df_fornecedor["Total"] * df_fornecedor["R$ Preço"]
                        df_exibicao = df_fornecedor[["Cód", "Produtos", "Total", "R$ Preço", "R$ Total"]].copy()

                        df_exibicao['Produtos'] = pd.Categorical(df_exibicao['Produtos'], categories=LISTA_NOMES_PRODUTOS)
                        altura_dinamica = int((len(df_exibicao) + 2) * 36) + 5
                        
                        col_cfg_forn = {
                            "Cód": st.column_config.NumberColumn(disabled=False, format="%d"),
                            "Produtos": st.column_config.SelectboxColumn("Produtos", options=LISTA_NOMES_PRODUTOS, disabled=False),
                            "Total": st.column_config.NumberColumn("Total", disabled=False, format="%d"),
                            "R$ Preço": st.column_config.NumberColumn("R$ Preço", format="R$ %.2f", disabled=False),
                            "R$ Total": st.column_config.NumberColumn("R$ Total", format="R$ %.2f", disabled=True)
                        }
                        
                        df_forn_edit = st.data_editor(
                            df_exibicao, 
                            hide_index=True, 
                            use_container_width=True, 
                            column_config=col_cfg_forn,
                            height=altura_dinamica,
                            num_rows="dynamic",
                            key=f"forn_{fornecedor}_{st.session_state['reset_counter']}"
                        )
                        
                        soma_dinamica = (pd.to_numeric(df_forn_edit["Total"], errors='coerce').fillna(0) * pd.to_numeric(df_forn_edit["R$ Preço"], errors='coerce').fillna(0)).sum()
                        
                        st.markdown(f"""
                            <div style="text-align:right; font-weight:700; margin-top:8px; color:var(--green-bright); font-size:16px;">
                                Total Final: R$ {soma_dinamica:,.2f}
                            </div>
                        """, unsafe_allow_html=True)
        st.write("<br>", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# ROTA 5: CATÁLOGO DE PRODUTOS
# ─────────────────────────────────────────────
elif perfil_navegacao == "Catálogo de Produtos":
    st.markdown("""
    <div class="page-header" style="background: linear-gradient(90deg, var(--green-dark) 0%, #0d2018 100%); padding: 14px 20px; border-radius: 10px; margin-bottom: 22px;">
        <span style="font-size: 26px; margin-right: 12px;">🏷️</span>
        <div style="display: inline-block; vertical-align: top;">
            <div style="font-size: 20px; font-weight: 700; color: var(--text-header);">Catálogo de Produtos</div>
            <div style="font-size: 12px; color: var(--text-muted); margin-top: 2px;">Gerencie itens, defina o setor (Box/Pedra) e as permissões por loja</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    with st.container(border=True):
        st.caption("➕ Adicione produtos na última linha  •  🗑️ Selecione a linha e pressione **Delete** para remover  •  ✅ Checkboxes controlam visibilidade por loja")

        config_catalogo = {
            "Código":    st.column_config.NumberColumn("Cód. Interno", width=90, required=True, min_value=0, format="%d"),
            "Descrição": st.column_config.TextColumn("Descrição do Item", width=310, required=True),
            "Tipo":      st.column_config.SelectboxColumn("Setor", options=["Box", "Pedra"], width=100, required=True),
        }
        for loja in LOJAS:
            config_catalogo[loja] = st.column_config.CheckboxColumn(loja, default=True, width=70)

        df_cat_editado = st.data_editor(
            df_produtos,
            num_rows="dynamic",
            column_config=config_catalogo,
            hide_index=True,
            use_container_width=True,
            height=580
        )

        st.divider()
        col_atualizar, col_info, _ = st.columns([2, 4, 4])
        with col_atualizar:
            if st.button("🔄 Atualizar Catálogo", type="primary", use_container_width=True):
                conn.update(worksheet="Produtos", data=df_cat_editado)
                st.cache_data.clear()
                st.success("✅ Catálogo e permissões atualizados para todas as lojas!")
                st.rerun()
        with col_info:
            total_prods = len(df_cat_editado)
            st.info(f"📦 **{total_prods}** produtos cadastrados  •  "
                    f"**{len(LOJAS)}** lojas configuradas")
